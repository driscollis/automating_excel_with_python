# creating_sheets.py

import openpyxl


def create_worksheets(path):
    workbook = openpyxl.Workbook()
    print(workbook.sheetnames)
    # Add a new worksheet
    workbook.create_sheet()
    print(workbook.sheetnames)
    # Insert a worksheet
    workbook.create_sheet(index=1, title="Second sheet")
    print(workbook.sheetnames)
    workbook.save(path)


if __name__ == "__main__":
    create_worksheets("sheets.xlsx")
