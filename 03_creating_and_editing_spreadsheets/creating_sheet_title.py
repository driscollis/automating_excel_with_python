# creating_sheet_title.py

from openpyxl import Workbook


def create_sheets(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hello"
    sheet2 = workbook.create_sheet(title="World")
    workbook.save(path)


if __name__ == "__main__":
    create_sheets("hello_sheets.xlsx")
