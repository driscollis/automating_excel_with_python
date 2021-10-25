# delete_sheets.py

import openpyxl


def create_worksheets(path):
    workbook = openpyxl.Workbook()
    workbook.create_sheet()
    print(workbook.sheetnames)
    # Insert a worksheet
    workbook.create_sheet(index=1, title="Second sheet")
    print(workbook.sheetnames)
    del workbook["Second sheet"]
    print(workbook.sheetnames)
    workbook.save(path)


if __name__ == "__main__":
    create_worksheets("del_sheets.xlsx")