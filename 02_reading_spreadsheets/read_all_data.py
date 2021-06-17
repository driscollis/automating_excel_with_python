# read_all_data.py

import openpyxl
from openpyxl import load_workbook


def read_all_data(path):
    workbook = load_workbook(filename=path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"{sheet.title=}")
        for row in sheet.rows:
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    # Skip this cell
                    continue

                print(f"{cell.column_letter}{cell.row} = {cell.value}")


if __name__ == "__main__":
    read_all_data("books.xlsx")