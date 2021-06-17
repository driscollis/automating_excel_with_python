# open_workbook.py

from openpyxl import load_workbook


def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f"Worksheet names: {workbook.sheetnames}")
    sheet = workbook.active
    print(sheet)
    print(f"The title of the Worksheet is: {sheet.title}")


if __name__ == "__main__":
    open_workbook("books.xlsx")
