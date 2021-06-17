# read_all_data_values.py

import openpyxl
from openpyxl import load_workbook


def read_all_data(path):
    workbook = load_workbook(filename=path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"{sheet.title=}")
        for value in sheet.iter_rows(values_only=True):
            print(value)


if __name__ == "__main__":
    read_all_data("books.xlsx")