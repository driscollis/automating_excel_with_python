# excel_to_csv.py

import csv
import openpyxl

from openpyxl import load_workbook


def excel_to_csv(excel_file, csv_file):
    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active
    csv_data = []

    # Read data from Excel
    for value in sheet.iter_rows(values_only=True):
        csv_data.append(list(value))

    # Write to CSV
    with open(csv_file, 'w') as csv_file_obj:
        writer = csv.writer(csv_file_obj, delimiter=',')
        for line in csv_data:
            writer.writerow(line)


if __name__ == "__main__":
    excel_to_csv("books.xlsx", "new_books.csv")